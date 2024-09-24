from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

from bs4 import BeautifulSoup

from openpyxl import Workbook
from openpyxl import load_workbook

import time
import re
from webdriver_manager.chrome import ChromeDriverManager

# Load the workbook and the active sheet
workbook = load_workbook("Warehouse_Ship_Days.xlsx")
sheet = workbook.active

# Initialize an empty array to store the extracted data
shipDates = []

# Loop through each row in the sheet
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
    first_col = row[0].value  # First column value
    second_col = row[2].value  # Second column value
    shipDates.append([first_col, second_col])

# Set up connection to an already-open Chrome instance with remote debugging
chrome_options = webdriver.ChromeOptions()
chrome_options.debugger_address = "localhost:9222"  # Connect to the existing Chrome session

# Initialize the WebDriver and connect to the existing Chrome
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

arnArr = [[]]

# Workbook setup for saving extracted data
workbook = Workbook()
sheet = workbook.active
sheet.title = 'ARN and ASN'
sheet.append(["ARN", "ASN", "Amazon Label", "UPS Tracking", "Warehouse Name"])


# Function to extract ARNs and ASNs from the page's BeautifulSoup object
def getARN(soup, sheet):
    for kat_label in soup.find_all("kat-label", class_="kat-label-light-text"):
        text_content = kat_label.get("text", "")
        # Check if the text contains the specific date
        if "Pickup: Fri, Sep 20, 2024 CDT" in text_content:
            # Extract the id and split by "-"
            label_id = kat_label.get("id", "")
            parts = label_id.split("-")
            if len(parts) >= 5:
                first_id = parts[3]
                second_id = parts[4]
                print(f"Extracted ARN: {first_id}, ASN: {second_id}")
                arnArr.append([first_id, second_id])

# Function to retrieve the shadow root of an element
def get_shadow_root(driver, css_selector):
    shadow_host = driver.find_element(By.CSS_SELECTOR, css_selector)  # Find the shadow host element
    shadow_root = driver.execute_script('return arguments[0].shadowRoot', shadow_host)  # Get the shadow root
    return shadow_root

# Main loop to scrape pages and navigate using the "Next" button
def main():
    driver.get('https://vendorcentral.amazon.com/kt/vendor/members/afi-shipment-mgr/shippingqueue')
    time.sleep(2)  # Wait for manual login
    #https://vendorcentral.amazon.com/kt/vendor/members/afi-shipment-mgr/shipmentdetail?rr=31650492961&asn=42169119692
    while True:
        # Get the page source after interaction and parse with BeautifulSoup
        soup = BeautifulSoup(driver.page_source, 'html.parser')

        # Extract ARNs and ASNs from the current page
        getARN(soup, sheet)
        if len(arnArr) > 30:
            break
        try:
            # Look for the "Next" button and click it using Selenium
            next_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//div[@id='sq-pag-next-div']//kat-label[@class='kat-label-link-text']//span[contains(text(), 'next >')]"))
            )
            print("Next button found. Clicking it...")
            driver.execute_script("arguments[0].click();", next_button)  # Use JavaScript to click the button
            time.sleep(3)  # Wait for the next page to load

        except TimeoutException:
            # If no next button is found, break the loop
            print("No more pages or 'Next' button not found. Exiting.")
            break

        except NoSuchElementException:
            print("No 'Next' button was found on this page.")
            break
    arnArr.pop(0)
    print(f"Array Size: {len(arnArr)}")
    for arn in arnArr:
        trckNumbers = []
        amznLbls = []
        wrhs = ""
        driver.get(f"https://vendorcentral.amazon.com/kt/vendor/members/afi-shipment-mgr/asnsubmission?arn={arn[0]}&asnId={arn[1]}")
        
        image = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//img[@height='45']"))
        )
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        #
        #Getting Warehouse Number
        # 
        # Regular expression to match 4 characters followed by a comma (XXXX,)
        pattern = re.compile(r'^[A-Za-z0-9]{4},')


        element = soup.findAll("kat-link", attrs={"slot": "trigger"})
        element = element[1]  
        if element:
            # Access the label attribute from the element
            label_value = element.get('label')
            print(label_value)
            if pattern.match(label_value):
                wrhs = label_value
                print(f"Matching span found: {label_value}")
        # Go to step 3
        kat_button = driver.execute_script("""
            return document.querySelector('kat-button[label="Continue to step 2"]');
        """)
        if kat_button:
            driver.execute_script("arguments[0].click();", kat_button)
        kat_button = driver.execute_script("""
            return document.querySelector('kat-button[label="Continue to step 3"]');
        """)
        if kat_button:
            driver.execute_script("arguments[0].click();", kat_button)            
        #
        #Getting Tracking Numbers
        # 
        tracking_number_cell = driver.find_elements(By.CSS_SELECTOR, 'div[col-id="carrierTrackingNumber"]')
        actions = ActionChains(driver)
        actions.double_click(tracking_number_cell[1]).perform()
        WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, '.ag-rich-select-row'))
        )
        rich_select_rows = driver.find_elements(By.CSS_SELECTOR, '.ag-rich-select-row')
        for row in rich_select_rows:
            if row.text is not None:
                trckNumbers.append(row.text)
        #
        #Getting Amazon Labels
        # 
        amazon_labels = driver.find_elements(By.CSS_SELECTOR, 'div[col-id="cartonLabelBarcode"]')
        for lbl in amazon_labels:
            if lbl.text.startswith("AMZN"):
                amznLbls.append(lbl.text)
        for i in range(len(amznLbls)):
            print(arn[0], arn[1], amznLbls[i], trckNumbers[i], wrhs)
            sheet.append([arn[0], arn[1], amznLbls[i], trckNumbers[i], wrhs])
            tracking_number_cell = driver.find_elements(By.CSS_SELECTOR, 'div[col-id="carrierTrackingNumber"]')
            actions = ActionChains(driver)
            actions.double_click(tracking_number_cell[i+1]).perform()
            rich_select_rows = driver.find_elements(By.CSS_SELECTOR, '.ag-rich-select-row')
            for row in rich_select_rows:
                if row.text == trckNumbers[i]:
                    row.click()
                    break
        foundDate = False
        for data in shipDates:
            if wrhs.split(',')[0] == data[0]:
                print(data[1])
                foundDate = True

        if foundDate:
            kat_button = driver.execute_script("""
                return document.querySelector('kat-button[label="Continue to step 4"]');
            """)
            if kat_button:
                driver.execute_script("arguments[0].click();", kat_button)           
            soup = BeautifulSoup(driver.page_source, 'html.parser') 
            
            shadow_root = get_shadow_root(driver, "kat-date-picker")
            elements = shadow_root.find_elements(By.CSS_SELECTOR, "*")
            print(elements[0].get_attribute('outerHTML'))
            # If this is the 'kat-input' element, you can change its value like this:
            if 'kat-input' in elements[0].get_attribute('outerHTML'):
                # Set a new value for the 'kat-input' field using JavaScript
                driver.execute_script("arguments[0].setAttribute('value', '09/30/2024');", elements[0])

        else:
            print(f"Couldn't find shipdate for warehouse: {wrhs}\nARN: {arn[0]}")
        
    # Save the workbook after all pages are processed
    workbook.save("ARN_ASN_Data.xlsx")
    print("Data saved to ARN_ASN_Data.xlsx")
    
# Run the main function
main()
